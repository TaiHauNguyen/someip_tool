"""Import the customer SOME/IP definition workbooks into the model.

Two workbook flavours are supported and both are read by *header name*, never
by fixed column index, because the customer sheets are not perfectly aligned
(e.g. PCU_Provider!Events has its data shifted two columns to the left of the
header row).  See MAPPING.md for the full Excel -> ARXML mapping.
"""

from __future__ import annotations

import copy
import os
import re
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

from someip_model import (
    EnumLiteral, EnumType, Event, EventGroup, EventParam, Endpoint,
    Service, StructMember, StructType, TsnStream, TsnSwitch, parse_int,
)


# --------------------------------------------------------------------------
# small helpers
# --------------------------------------------------------------------------
def _s(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _norm(v: Any) -> str:
    """Normalise a header cell for matching: lower case, no spaces/slashes."""
    return re.sub(r"[^a-z0-9]", "", _s(v).lower())


def camel_to_upper_snake(name: str) -> str:
    s = re.sub(r"(?<=[a-z0-9])(?=[A-Z])|(?<=[A-Z])(?=[A-Z][a-z])", "_", name)
    return re.sub(r"_+", "_", s).upper()


def default_vt(enum_name: str, literal: str) -> str:
    prefix = enum_name[:-4] if enum_name.endswith("Type") else enum_name
    return "%s_%s" % (camel_to_upper_snake(prefix), camel_to_upper_snake(literal))


# --------------------------------------------------------------------------
# Configuration sheet
# --------------------------------------------------------------------------
def read_configuration(ws) -> List[Tuple[str, str, str]]:
    """Return [(block, key, value)] with the block name carried down."""
    out: List[Tuple[str, str, str]] = []
    block = ""
    for row in ws.iter_rows(min_row=2, values_only=True):
        a = _s(row[0] if len(row) > 0 else "")
        b = _s(row[1] if len(row) > 1 else "")
        c = _s(row[2] if len(row) > 2 else "")
        if not a and not b and not c:
            continue
        if a:
            block = a
        out.append((block, b, c))
    return out


def _cfg(rows, block: str, key: str = "", default: str = "") -> str:
    nb, nk = _norm(block), _norm(key)
    for b, k, v in rows:
        if _norm(b) == nb and _norm(k) == nk:
            return v
    return default


def _has_block(rows, block: str) -> bool:
    nb = _norm(block)
    return any(_norm(b) == nb for b, _, _ in rows)


# --------------------------------------------------------------------------
# generic table reader
# --------------------------------------------------------------------------
def find_header(ws, required: List[str], max_scan: int = 12) -> Optional[Tuple[int, Dict[str, int]]]:
    """Find the first row holding every `required` header; return (row, {norm-header: col})."""
    req = [_norm(r) for r in required]
    for r in range(1, min(ws.max_row, max_scan) + 1):
        cols: Dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            h = _norm(ws.cell(row=r, column=c).value)
            if h and h not in cols:
                cols[h] = c
        if all(x in cols for x in req):
            return r, cols
    return None


def find_header_groups(ws, required: List[str], max_scan: int = 12) -> List[Tuple[int, Dict[str, int]]]:
    """Like find_header but returns every contiguous column block that matches.

    The DataStructures sheet holds the enum table and the struct table side by
    side, sometimes with their header rows on different lines.
    """
    req = [_norm(r) for r in required]
    found: List[Tuple[int, Dict[str, int]]] = []
    for r in range(1, min(ws.max_row, max_scan) + 1):
        # split the row into contiguous groups of non-empty header cells
        group: Dict[str, int] = {}
        groups: List[Dict[str, int]] = []
        for c in range(1, ws.max_column + 2):
            h = _norm(ws.cell(row=r, column=c).value) if c <= ws.max_column else ""
            if h:
                group.setdefault(h, c)
            elif group:
                groups.append(group)
                group = {}
        for g in groups:
            if all(x in g for x in req):
                found.append((r, g))
    return found


# --------------------------------------------------------------------------
# Events sheet
# --------------------------------------------------------------------------
EVENT_HEADERS = ["Index", "Name", "EventId", "PayloadLengthBytes"]


def read_events(ws) -> List[Event]:
    hit = find_header(ws, EVENT_HEADERS)
    if not hit:
        return []
    hrow, cols = hit

    def col(name: str) -> int:
        return cols.get(_norm(name), 0)

    c_transport = col("TransportProtocol")
    tail_start = col("MaximumSegmentLength") or c_transport

    events: List[Event] = []
    current: Optional[Event] = None
    for r in range(hrow + 1, ws.max_row + 1):
        def val(name: str, shift: int = 0) -> str:
            c = col(name)
            return _s(ws.cell(row=r, column=c + shift).value) if c else ""

        name = val("Name")
        # Detect a horizontal shift of the data relative to the header row by
        # locating the UDP/TCP cell, which is unambiguous.
        shift = 0
        if c_transport:
            for c in range(1, ws.max_column + 1):
                if _s(ws.cell(row=r, column=c).value).upper() in ("UDP", "TCP"):
                    shift = c - c_transport
                    break

        def tail(nm: str) -> str:
            """Read a column of the shifted tail region.

            A negative shift means the sheet simply omits the leading tail
            columns (MaximumSegmentLength / SeparationTime); those must read as
            empty instead of wrapping back onto EventId / PayloadLengthBytes.
            """
            c = col(nm)
            if not c:
                return ""
            if c < tail_start:
                return _s(ws.cell(row=r, column=c).value)
            if c + shift < tail_start:
                return ""
            return _s(ws.cell(row=r, column=c + shift).value)

        if name:
            current = Event(
                index=parse_int(val("Index"), len(events) + 1),
                name=name,
                event_id="0x%04X" % parse_int(val("EventId")),
                payload_length=parse_int(val("PayloadLengthBytes")),
                max_segment_length=tail("MaximumSegmentLength"),
                separation_time=tail("SeparationTime"),
                serializer=tail("Serializer"),
                transport=tail("TransportProtocol") or "UDP",
                event_group=tail("EventGroup"),
            )
            events.append(current)
        if current is None:
            continue
        pname = tail("ParameterName")
        if pname:
            current.params.append(EventParam(
                index=parse_int(tail("ParameterIndex"), len(current.params) + 1),
                name=pname,
                type=tail("ParameterType"),
                description=tail("ParameterDescription"),
            ))
    return events


# --------------------------------------------------------------------------
# EventGroups sheet
# --------------------------------------------------------------------------
def read_event_groups(ws) -> List[EventGroup]:
    hit = find_header(ws, ["Index", "Name", "EventGroupId"])
    if not hit:
        return []
    hrow, cols = hit

    def val(r: int, name: str) -> str:
        c = cols.get(_norm(name), 0)
        return _s(ws.cell(row=r, column=c).value) if c else ""

    groups: List[EventGroup] = []
    for r in range(hrow + 1, ws.max_row + 1):
        if not val(r, "Name"):
            continue
        groups.append(EventGroup(
            index=parse_int(val(r, "Index"), len(groups) + 1),
            name=val(r, "Name"),
            group_id="0x%02X" % parse_int(val(r, "EventGroupId")),
            dest_zone=val(r, "DestinationZone"),
            dest_ipv4=val(r, "DestinationIpv4Address"),
            dest_mac=val(r, "DestinationMacAddress"),
            dest_udp_port=parse_int(val(r, "DestinationUdpPort")),
            transport=val(r, "TransportProtocol") or "UDP",
            routing_mode=val(r, "RoutingMode") or "StaticUnicast",
        ))
    return groups


# --------------------------------------------------------------------------
# DataStructures sheet
# --------------------------------------------------------------------------
LITERAL_HEADERS = ("enumerateliteral", "enumeral", "literal", "enumerate", "enum")


def read_data_structures(ws) -> Tuple[List[StructType], List[EnumType]]:
    structs = _read_structs(ws)
    enums = _read_enums(ws)
    return structs, enums


def _read_structs(ws) -> List[StructType]:
    hits = find_header_groups(ws, ["Index", "Name", "Element", "Type"])
    structs: List[StructType] = []
    for hrow, cols in hits:
        cur: Optional[StructType] = None
        for r in range(hrow + 1, ws.max_row + 1):
            def val(nm: str) -> str:
                c = cols.get(_norm(nm), 0)
                return _s(ws.cell(row=r, column=c).value) if c else ""

            nm = val("Name")
            if nm:
                cur = StructType(name=nm)
                structs.append(cur)
            if cur is None:
                continue
            el = val("Element")
            if el:
                cur.members.append(StructMember(
                    name=el, type=val("Type"), description=val("Description")))
    return structs


def _read_enums(ws) -> List[EnumType]:
    enums: List[EnumType] = []
    # locate the enum header row/columns: Index + Name + Type + a literal column
    for r in range(1, min(ws.max_row, 12) + 1):
        cols: Dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            h = _norm(ws.cell(row=r, column=c).value)
            if h and h not in cols:
                cols[h] = c
        lit_key = next((k for k in LITERAL_HEADERS if k in cols), None)
        if not lit_key or "index" not in cols or "name" not in cols or "type" not in cols:
            continue
        if "element" in cols and cols["element"] < cols[lit_key]:
            continue  # this is the struct block, not the enum block
        c_lit = cols[lit_key]
        c_val = cols.get("value", 0)
        c_desc = cols.get("description", 0)
        # keep only the columns that belong to the enum block (left of the literal col + value)
        c_index, c_name, c_type = cols["index"], cols["name"], cols["type"]
        if c_index > c_lit:
            continue
        cur: Optional[EnumType] = None
        for rr in range(r + 1, ws.max_row + 1):
            def cell(c: int) -> str:
                return _s(ws.cell(row=rr, column=c).value) if c else ""

            nm = cell(c_name)
            if nm and cell(c_index):
                cur = EnumType(name=nm, base_type=cell(c_type) or "uint8_t")
                enums.append(cur)
            lit = cell(c_lit)
            if cur is not None and lit:
                cur.literals.append(EnumLiteral(
                    name=lit, value=parse_int(cell(c_val)),
                    description=cell(c_desc) if c_desc else ""))
        break
    _assign_vt(enums)
    return enums


def _assign_vt(enums: List[EnumType]) -> None:
    for en in enums:
        used: Dict[str, int] = {}
        for lit in en.literals:
            if lit.vt:
                continue
            base = default_vt(en.name, lit.name)
            n = used.get(base, 0) + 1
            used[base] = n
            lit.vt = base if n == 1 else "%s_%d" % (base, n)


# --------------------------------------------------------------------------
# Workbook -> Service
# --------------------------------------------------------------------------
def import_service(path: str) -> Service:
    wb = openpyxl.load_workbook(path, data_only=True)
    rows = read_configuration(wb["Configuration"]) if "Configuration" in wb.sheetnames else []

    provider = _has_block(rows, "ProvidedSomeipServiceInstance")
    role = "provider" if provider else "consumer"
    inst_block = "ProvidedSomeipServiceInstance" if provider else "ConsumedSomeipServiceInstance"

    instance_name = _cfg(rows, "ServiceInstanceName")
    tag = instance_name[:-7] if instance_name.endswith("Service") else instance_name

    svc = Service(
        role=role,
        tag=tag,
        instance_name=instance_name,
        interface_name=_cfg(rows, "ServiceInterface"),
        instance_id="0x%04X" % parse_int(_cfg(rows, "ServiceInstanceId")),
        major_version=parse_int(_cfg(rows, "Someip Interface Major Version"), 1),
        minor_version=parse_int(_cfg(rows, "Someip Interface Minor Version"), 0),
        sd_udp_port=parse_int(_cfg(rows, "SomeipServiceDiscovery", "UdpPort"), 30490),
        local=Endpoint(
            zone=_cfg(rows, "LocalMcuEndpoint", "Zone"),
            ipv4=_cfg(rows, "LocalMcuEndpoint", "Ipv4Address"),
            mac=_cfg(rows, "LocalMcuEndpoint", "MacAddress"),
        ),
        tsn_switch=TsnSwitch(
            model=_cfg(rows, "LocalTsnSwitch", "Model"),
            bridge_mac=_cfg(rows, "LocalTsnSwitch", "BridgeMacAddress"),
            ring_port_cw_neighbor=_cfg(rows, "LocalTsnSwitch", "RingPortCwNeighbor"),
            ring_port_ccw_neighbor=_cfg(rows, "LocalTsnSwitch", "RingPortCcwNeighbor"),
        ),
        tsn=TsnStream(
            zone=_cfg(rows, "TSN", "ZONE"),
            topology=_cfg(rows, "TSN", "Topology") or "Ring",
            vlan_id=parse_int(_cfg(rows, "TSN", "VLAN ID"), 5),
            vlan_priority=parse_int(_cfg(rows, "TSN", "VLAN Priority"), 6),
            traffic_class=_cfg(rows, "TSN", "TrafficClass"),
            traffic_profile=_cfg(rows, "TSN", "TrafficProfile"),
        ),
        deployment_name=_cfg(rows, "SomeipServiceDeployment", "Name"),
        interface_id="0x%04X" % parse_int(_cfg(rows, "SomeipServiceDeployment", "ServiceInterfaceId")),
        service_instance_name=_cfg(rows, inst_block, "Name"),
        load_balancing_priority=_cfg(rows, inst_block, "LoadBalancingPriority", "-"),
        load_balancing_weight=_cfg(rows, inst_block, "LoadBalancingWeight", "-"),
        mapping_name=_cfg(rows, "SomeipServiceInstanceToMachineMapping", "Name"),
        communication_connector=_cfg(rows, "SomeipServiceInstanceToMachineMapping", "CommunicationConnector", "-"),
        secoc_com_props_multicast=_cfg(rows, "SomeipServiceInstanceToMachineMapping", "SecOcComPropsForMulticasts", "-"),
        secure_com_props_tcp=_cfg(rows, "SomeipServiceInstanceToMachineMapping", "SecureComPropsForTcps", "-"),
        secure_com_props_udp=_cfg(rows, "SomeipServiceInstanceToMachineMapping", "SecureComPropsForUdps", "-"),
        tcp_port=_cfg(rows, "SomeipServiceInstanceToMachineMapping", "TcpPort", "-"),
        udp_port=parse_int(_cfg(rows, "SomeipServiceInstanceToMachineMapping", "UdpPort")),
        udp_collection_buffer_size_threshold=_cfg(
            rows, "SomeipServiceInstanceToMachineMapping", "UdpCollectionBufferSizeThreshold", "-"),
        routing_mode=_cfg(rows, "StaticEventRouting", "Mode") or "StaticUnicast",
        source_file=os.path.basename(path),
    )

    if "Events" in wb.sheetnames:
        svc.events = read_events(wb["Events"])
    if "EventGroups" in wb.sheetnames:
        svc.event_groups = read_event_groups(wb["EventGroups"])
    if "DataStructures" in wb.sheetnames:
        svc.structs, svc.enums = read_data_structures(wb["DataStructures"])

    # A consumer receives on its own endpoint; the remote provider endpoint is
    # not part of the customer workbook, so seed it from the ring neighbour.
    if not svc.is_provider:
        svc.remote = Endpoint(zone=svc.tsn_switch.ring_port_ccw_neighbor or "REMOTE")
    return svc


def import_project(paths: List[str], project=None, log: Optional[List[str]] = None):
    """Import one or more workbooks.

    `project` is the *base*: an ARXML or JSON you already have.  The workbooks
    win for everything they define; the base supplies what a workbook cannot
    express (a consumer's remote provider endpoint, the type names and <VT>
    texts the ECU code already uses).  See resolve.py for those rules.
    """
    import resolve
    from someip_model import Project

    base = project
    prj = Project()
    if base is not None:
        prj = copy.deepcopy(base)
        prj.services = [s for s in prj.services]

    for p in paths:
        svc = import_service(p)
        old = prj.find_service(svc.tag)
        if old is not None:
            prj.services[prj.services.index(old)] = svc
        else:
            prj.services.append(svc)

    messages = resolve.apply_all(prj, base)
    if log is not None:
        log.extend(messages)
    _sync_project(prj)
    return prj


def _sync_project(prj) -> None:
    """Propagate service level settings that are really project wide."""
    for s in prj.services:
        if s.tsn.vlan_id:
            prj.vlan_id = s.tsn.vlan_id
        if s.tsn.vlan_priority:
            prj.vlan_priority = s.tsn.vlan_priority
        break
